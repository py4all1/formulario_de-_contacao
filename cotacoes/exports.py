"""Geração de exportações da cotação: JSON, Excel (.xlsx) e PDF."""
import io
import json
from decimal import Decimal

from django.http import HttpResponse
from django.utils import timezone


def _dec(v):
    return float(v) if isinstance(v, Decimal) else v


# ---------------------------------------------------------------------------
# Estrutura de dados serializável
# ---------------------------------------------------------------------------
def cotacao_to_dict(cotacao):
    return {
        'id': cotacao.id,
        'status': cotacao.get_status_display(),
        'criado_em': cotacao.criado_em.isoformat(),
        'electrolux': {
            'filial': cotacao.filial.planta,
            'cnpj_filial': cotacao.filial.cnpj_formatado,
            'planta': cotacao.filial.planta,
            'municipio': cotacao.filial.municipio,
            'uf': cotacao.filial.uf,
            'cep': cotacao.filial.cep,
            'endereco': cotacao.filial.endereco,
            'tipo_fornecimento': cotacao.tipo_fornecimento,
        },
        'fornecedor': {
            'cnpj': cotacao.cnpj_formatado,
            'razao_social': cotacao.razao_social,
            'nome_fantasia': cotacao.nome_fantasia,
            'regime': cotacao.regime,
            'endereco': cotacao.endereco,
            'bairro': cotacao.bairro,
            'cidade': cotacao.cidade,
            'estado': cotacao.estado,
            'cep': cotacao.cep,
            'email': cotacao.email,
            'telefone': cotacao.telefone,
            'numero_proposta': cotacao.numero_proposta,
        },
        'condicoes': {
            'frete': cotacao.frete,
            'prazo_pagamento': cotacao.prazo_pagamento,
            'data_emissao': cotacao.data_emissao.isoformat() if cotacao.data_emissao else None,
        },
        'servicos': [{
            'valor_servico': _dec(i.valor_servico),
            'codigo_servico': i.codigo_servico,
            'descricao': i.descricao,
            'perc_iss': _dec(i.perc_iss),
            'perc_inss': _dec(i.perc_inss),
            'perc_csrf': _dec(i.perc_csrf),
            'perc_irrf': _dec(i.perc_irrf),
            'total_retido': _dec(i.total_retido),
            'codigo_sap': i.codigo_sap,
            'valor_liquido': _dec(i.valor_liquido),
        } for i in cotacao.servicos.all()],
        'mercadorias': [{
            'item': i.item,
            'quantidade': _dec(i.quantidade),
            'unidade': i.unidade,
            'prazo_entrega': i.prazo_entrega,
            'valor_unitario': _dec(i.valor_unitario),
            'cst_icms': i.cst_icms,
            'cfop': i.cfop,
            'cod_cest': i.cod_cest,
            'categoria': i.categoria,
            'ncm': i.ncm,
            'excecao': i.excecao,
            'perc_ipi': _dec(i.perc_ipi),
            'valor_ipi': _dec(i.valor_ipi),
            'perc_icms': _dec(i.perc_icms),
            'perc_reducao_base': _dec(i.perc_reducao_base),
            'base_icms': _dec(i.base_icms),
            'valor_icms': _dec(i.valor_icms),
            'icms_st': _dec(i.icms_st),
            'valor_total': _dec(i.valor_total),
            'observacoes': i.observacoes,
        } for i in cotacao.mercadorias.all()],
        'aceite_termos': {
            'aceito': cotacao.aceite_termos,
            'data': cotacao.aceite_em.isoformat() if cotacao.aceite_em else None,
            'ip': cotacao.aceite_ip,
        },
        'totais': {
            'total_servico': _dec(cotacao.total_servico),
            'total_retido': _dec(cotacao.total_retido),
            'total_material': _dec(cotacao.total_material),
            'total_ipi': _dec(cotacao.total_ipi),
            'total_icms': _dec(cotacao.total_icms),
            'valor_liquido': _dec(cotacao.valor_liquido),
        },
    }


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------
def export_json(cotacao):
    payload = json.dumps(cotacao_to_dict(cotacao), ensure_ascii=False, indent=2)
    resp = HttpResponse(payload, content_type='application/json; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="cotacao_{cotacao.id}.json"'
    return resp


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def export_excel(cotacao):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    azul = PatternFill('solid', fgColor='1F4E78')
    cinza = PatternFill('solid', fgColor='E7E6E6')
    branco = Font(color='FFFFFF', bold=True, size=12)
    bold = Font(bold=True)
    thin = Side(style='thin', color='BFBFBF')
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    def titulo(ws, cell, texto):
        ws[cell] = texto
        ws[cell].fill = azul
        ws[cell].font = branco

    # --- Aba Resumo --------------------------------------------------------
    ws = wb.active
    ws.title = 'Cotação'
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 55
    titulo(ws, 'A1', 'COTAÇÃO / CADASTRO DE FORNECEDOR')
    ws.merge_cells('A1:B1')

    pares = [
        ('DADOS DA ELECTROLUX', ''),
        ('Filial', cotacao.filial.planta),
        ('CNPJ filial', cotacao.filial.cnpj_formatado),
        ('Planta / UF', f'{cotacao.filial.municipio}/{cotacao.filial.uf}'),
        ('Endereço', cotacao.filial.endereco),
        ('CEP', cotacao.filial.cep),
        ('Tipo de fornecimento', cotacao.tipo_fornecimento),
        ('DADOS DO FORNECEDOR', ''),
        ('CNPJ', cotacao.cnpj_formatado),
        ('Razão social', cotacao.razao_social),
        ('Nome fantasia', cotacao.nome_fantasia),
        ('Regime', cotacao.regime),
        ('Endereço', cotacao.endereco),
        ('Bairro', cotacao.bairro),
        ('Cidade / UF', f'{cotacao.cidade}/{cotacao.estado}'),
        ('CEP', cotacao.cep),
        ('E-mail', cotacao.email),
        ('Telefone / WhatsApp', cotacao.telefone),
        ('Número da proposta', cotacao.numero_proposta),
        ('CONDIÇÕES', ''),
        ('Frete', cotacao.frete),
        ('Prazo de pagamento', cotacao.prazo_pagamento),
        ('Data de emissão', cotacao.data_emissao.strftime('%d/%m/%Y') if cotacao.data_emissao else ''),
        ('TOTAIS', ''),
        ('Total serviço (líquido)', _dec(cotacao.total_servico)),
        ('Total retido', _dec(cotacao.total_retido)),
        ('Total material', _dec(cotacao.total_material)),
        ('Total IPI', _dec(cotacao.total_ipi)),
        ('Total ICMS', _dec(cotacao.total_icms)),
        ('Valor líquido a receber', _dec(cotacao.valor_liquido)),
    ]
    r = 3
    for k, v in pares:
        if v == '' and k.isupper():
            ws.merge_cells(f'A{r}:B{r}')
            ws[f'A{r}'] = k
            ws[f'A{r}'].fill = cinza
            ws[f'A{r}'].font = bold
        else:
            ws[f'A{r}'] = k
            ws[f'A{r}'].font = bold
            ws[f'B{r}'] = v
        r += 1

    # --- Aba Serviços ------------------------------------------------------
    if cotacao.servicos.exists():
        wss = wb.create_sheet('Serviços')
        head = ['Valor', 'Código', 'Descrição', '% ISS', '% INSS', '% CSRF',
                '% IRRF', 'Total retido', 'Código SAP', 'Valor líquido']
        for c, h in enumerate(head, 1):
            cell = wss.cell(1, c, h)
            cell.fill = azul; cell.font = branco; cell.border = borda
        for ri, i in enumerate(cotacao.servicos.all(), 2):
            vals = [_dec(i.valor_servico), i.codigo_servico, i.descricao,
                    _dec(i.perc_iss), _dec(i.perc_inss), _dec(i.perc_csrf),
                    _dec(i.perc_irrf), _dec(i.total_retido), i.codigo_sap,
                    _dec(i.valor_liquido)]
            for c, v in enumerate(vals, 1):
                cell = wss.cell(ri, c, v); cell.border = borda
        for col, w in zip('ABCDEFGHIJ', [12, 10, 45, 8, 8, 8, 8, 14, 16, 14]):
            wss.column_dimensions[col].width = w

    # --- Aba Mercadorias ---------------------------------------------------
    if cotacao.mercadorias.exists():
        wsm = wb.create_sheet('Mercadorias')
        head = ['Item', 'Qtd', 'Unid', 'Prazo', 'Vlr unit', 'CST', 'CFOP', 'CEST',
                'Categoria', 'NCM', 'Exceção', 'IPI %', 'Vlr IPI', 'ICMS %',
                '% Red base', 'Base ICMS', 'Vlr ICMS', 'ICMS ST', 'Vlr total', 'Obs']
        for c, h in enumerate(head, 1):
            cell = wsm.cell(1, c, h)
            cell.fill = azul; cell.font = branco; cell.border = borda
        for ri, i in enumerate(cotacao.mercadorias.all(), 2):
            vals = [i.item, _dec(i.quantidade), i.unidade, i.prazo_entrega,
                    _dec(i.valor_unitario), i.cst_icms, i.cfop, i.cod_cest,
                    i.categoria, i.ncm, i.excecao, _dec(i.perc_ipi), _dec(i.valor_ipi),
                    _dec(i.perc_icms), _dec(i.perc_reducao_base), _dec(i.base_icms),
                    _dec(i.valor_icms), _dec(i.icms_st), _dec(i.valor_total), i.observacoes]
            for c, v in enumerate(vals, 1):
                cell = wsm.cell(ri, c, v); cell.border = borda
        wsm.column_dimensions['A'].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="cotacao_{cotacao.id}.xlsx"'
    return resp


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------
def export_pdf(cotacao):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                    Spacer)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=15 * mm, bottomMargin=15 * mm,
                            leftMargin=14 * mm, rightMargin=14 * mm,
                            title=f'Cotação {cotacao.id}')
    styles = getSampleStyleSheet()
    h = ParagraphStyle('h', parent=styles['Heading2'], textColor=colors.HexColor('#1F4E78'))
    small = ParagraphStyle('s', parent=styles['Normal'], fontSize=8)
    AZUL = colors.HexColor('#1F4E78')
    CINZA = colors.HexColor('#E7E6E6')
    elements = []

    elements.append(Paragraph('Cotação / Cadastro de Fornecedor', styles['Title']))
    elements.append(Paragraph(
        f'#{cotacao.id} — emitido em {timezone.localtime(cotacao.criado_em).strftime("%d/%m/%Y %H:%M")}',
        small))
    elements.append(Spacer(1, 6 * mm))

    def bloco(titulo, linhas):
        elements.append(Paragraph(titulo, h))
        data = [[Paragraph(f'<b>{k}</b>', small), Paragraph(str(v or '—'), small)]
                for k, v in linhas]
        t = Table(data, colWidths=[55 * mm, 120 * mm])
        t.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#BFBFBF')),
            ('BACKGROUND', (0, 0), (0, -1), CINZA),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 4 * mm))

    bloco('Dados da Electrolux', [
        ('Filial', cotacao.filial.planta),
        ('CNPJ filial', cotacao.filial.cnpj_formatado),
        ('Local', f'{cotacao.filial.municipio}/{cotacao.filial.uf} — CEP {cotacao.filial.cep}'),
        ('Endereço', cotacao.filial.endereco),
        ('Tipo de fornecimento', cotacao.tipo_fornecimento),
    ])
    bloco('Dados do Fornecedor', [
        ('CNPJ', cotacao.cnpj_formatado),
        ('Razão social', cotacao.razao_social),
        ('Nome fantasia', cotacao.nome_fantasia),
        ('Regime', cotacao.regime),
        ('Endereço', f'{cotacao.endereco}, {cotacao.bairro} — {cotacao.cidade}/{cotacao.estado} — CEP {cotacao.cep}'),
        ('Contato', f'{cotacao.email} · {cotacao.telefone}'),
        ('Proposta nº', cotacao.numero_proposta),
        ('Frete / Pagamento', f'{cotacao.frete} · {cotacao.prazo_pagamento}'),
    ])

    def tabela(titulo, head, rows, widths):
        elements.append(Paragraph(titulo, h))
        data = [head] + rows
        t = Table(data, colWidths=widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), AZUL),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#BFBFBF')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 4 * mm))

    def br(v):
        return f'{v:,.2f}'.replace(',', '#').replace('.', ',').replace('#', '.')

    if cotacao.servicos.exists():
        rows = [[br(i.valor_servico), i.codigo_servico,
                 Paragraph((i.descricao or '')[:60], small), br(i.total_retido),
                 i.codigo_sap, br(i.valor_liquido)]
                for i in cotacao.servicos.all()]
        tabela('Serviços',
               ['Valor', 'Código', 'Descrição', 'Retido', 'SAP', 'Líquido'],
               rows, [22 * mm, 16 * mm, 70 * mm, 22 * mm, 24 * mm, 22 * mm])

    if cotacao.mercadorias.exists():
        rows = [[Paragraph((i.item or '')[:40], small), br(i.quantidade), i.unidade,
                 br(i.valor_unitario), br(i.valor_ipi), br(i.valor_icms), br(i.valor_total)]
                for i in cotacao.mercadorias.all()]
        tabela('Mercadorias',
               ['Item', 'Qtd', 'Un', 'Vlr unit', 'IPI', 'ICMS', 'Total'],
               rows, [55 * mm, 18 * mm, 12 * mm, 24 * mm, 22 * mm, 22 * mm, 24 * mm])

    elements.append(Paragraph('Totais', h))
    tot = [
        ['Total serviço (líquido)', br(cotacao.total_servico)],
        ['Total retido', br(cotacao.total_retido)],
        ['Total material', br(cotacao.total_material)],
        ['Total IPI', br(cotacao.total_ipi)],
        ['Total ICMS', br(cotacao.total_icms)],
        ['Valor líquido a receber', br(cotacao.valor_liquido)],
    ]
    t = Table(tot, colWidths=[120 * mm, 55 * mm])
    t.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#BFBFBF')),
        ('BACKGROUND', (0, -1), (-1, -1), AZUL),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    elements.append(t)

    doc.build(elements)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="cotacao_{cotacao.id}.pdf"'
    return resp
